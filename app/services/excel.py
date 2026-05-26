import io
import json
import csv
import os
import re
import pandas as pd
from typing import List, Dict, Any, Optional
import logging
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

from app.utils.exceptions import FileProcessingError

logger = logging.getLogger(__name__)


class ExcelService:
    """Service for converting CSV data to XLSX format."""
    
    def __init__(self):
        self.header_font = Font(bold=True, color="FFFFFF")
        self.header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        self.alignment = Alignment(horizontal="left", vertical="top")
    
    def csv_to_xlsx(self, csv_content: str, sheet_name: str = "Sheet1") -> bytes:
        """
        Convert CSV content to XLSX format.
        
        Args:
            csv_content: CSV data as string
            sheet_name: Name for the Excel sheet
            
        Returns:
            XLSX file as bytes
            
        Raises:
            FileProcessingError: If CSV parsing fails
        """
        try:
            # Log the CSV content for debugging
            logger.info(f"Converting CSV to XLSX. Content length: {len(csv_content)}")
            logger.info(f"CSV content preview (first 200 chars): {csv_content[:200]}")
            
            # Handle empty or whitespace-only content
            if not csv_content or not csv_content.strip():
                logger.warning("CSV content is empty or whitespace only")
                # Create a minimal Excel file with a message
                return self._create_empty_xlsx_with_message("No data extracted from image", sheet_name)
            
            # Parse CSV content into DataFrame
            df = pd.read_csv(io.StringIO(csv_content))
            
            # Check if DataFrame is empty
            if df.empty:
                logger.warning("Parsed DataFrame is empty")
                return self._create_empty_xlsx_with_message("No tabular data found in image", sheet_name)
            
            # Handle empty cells
            df = df.fillna("")
            
            # Convert to XLSX
            xlsx_bytes = self._dataframe_to_xlsx_bytes(df, sheet_name)
            
            logger.info(f"Successfully converted CSV to XLSX: {len(df)} rows, {len(df.columns)} columns")
            return xlsx_bytes
            
        except pd.errors.EmptyDataError:
            logger.warning("CSV EmptyDataError - creating empty Excel file")
            return self._create_empty_xlsx_with_message("No data extracted from image", sheet_name)
        except pd.errors.ParserError as e:
            logger.warning(f"CSV ParserError: {str(e)} - attempting to create raw text Excel")
            return self._create_raw_text_xlsx(csv_content, sheet_name)
        except Exception as e:
            logger.error(f"Excel conversion error: {str(e)}")
            # Instead of raising an error, create an Excel file with the error message
            return self._create_empty_xlsx_with_message(f"Error processing image: {str(e)}", sheet_name)

    def table_to_csv(self, csv_content: str) -> bytes:
        """Return raw detected table rows as a downloadable UTF-8 CSV."""
        return (csv_content or "").encode("utf-8-sig")
    
    def create_batch_xlsx(
        self, 
        batch_results: List[Dict[str, Any]], 
        output_type: str = "separate"
    ) -> bytes:
        """
        Create XLSX file(s) from batch processing results.
        
        Args:
            batch_results: List of batch processing results
            output_type: "separate", "consolidated", or "concatenated"
            
        Returns:
            XLSX file as bytes
        """
        if output_type == "consolidated":
            return self._create_consolidated_xlsx(batch_results)
        elif output_type == "concatenated":
            return self._create_concatenated_xlsx(batch_results)
        else:  # separate
            return self._create_separate_xlsx(batch_results)
    
    def _create_consolidated_xlsx(self, batch_results: List[Dict[str, Any]]) -> bytes:
        """
        Create a single XLSX file with multiple sheets.
        
        Args:
            batch_results: List of batch processing results
            
        Returns:
            XLSX file as bytes
        """
        try:
            workbook = Workbook()
            # Remove default sheet
            workbook.remove(workbook.active)
            
            successful_sheets = 0
            
            for i, result in enumerate(batch_results):
                sheet_name = f"Image_{i+1}"
                worksheet = workbook.create_sheet(title=sheet_name)
                
                if result["success"] and result["data"]:
                    try:
                        # Log the data being processed
                        logger.info(f"Processing sheet {sheet_name}, data length: {len(result['data'])}")
                        logger.info(f"Data preview: {result['data'][:200]}...")
                        
                        df = pd.read_csv(io.StringIO(result["data"]))
                        df = df.fillna("")
                        
                        if not df.empty:
                            # Add data to worksheet
                            for r in dataframe_to_rows(df, index=False, header=True):
                                worksheet.append(r)
                            
                            # Style the header row
                            self._style_header_row(worksheet)
                            successful_sheets += 1
                            logger.info(f"Successfully added sheet {sheet_name} with {len(df)} rows")
                        else:
                            # Empty dataframe
                            worksheet.append(["Status", "No data extracted from image"])
                            worksheet.append(["Raw Data", result["data"] if result["data"] else "No raw data"])
                        
                    except Exception as e:
                        logger.warning(f"Failed to parse CSV for sheet {sheet_name}: {str(e)}")
                        logger.warning(f"Raw data that failed: {result['data'][:500]}...")
                        # Add error information to sheet
                        worksheet.append(["Error", f"Failed to parse data: {str(e)}"])
                        worksheet.append(["Raw Data", result["data"] if result["data"] else "No data"])
                        if result["data"]:
                            # Split raw data into lines and add them
                            lines = result["data"].split('\n')[:10]  # First 10 lines
                            for j, line in enumerate(lines):
                                worksheet.append([f"Line {j+1}", line])
                else:
                    # Failed result
                    error_msg = result.get("error", "Unknown error")
                    worksheet.append(["Status", "Processing Failed"])
                    worksheet.append(["Error", error_msg])
                    logger.warning(f"Sheet {sheet_name} failed: {error_msg}")
            
            # Add a summary sheet
            summary_sheet = workbook.create_sheet(title="Summary", index=0)
            summary_sheet.append(["Batch Processing Summary"])
            summary_sheet.append(["Total Images", str(len(batch_results))])
            
            successful_count = sum(1 for r in batch_results if r["success"])
            summary_sheet.append(["Successfully Processed", str(successful_count)])
            summary_sheet.append(["Failed Processing", str(len(batch_results) - successful_count)])
            summary_sheet.append(["Sheets with Valid Data", str(successful_sheets)])
            
            # Style the summary sheet
            for row in summary_sheet['A1:B5']:
                for cell in row:
                    if cell.column == 1:  # First column
                        cell.font = self.header_font
                        cell.fill = self.header_fill
            
            # If no successful sheets were added, add debugging info
            if successful_sheets == 0:
                summary_sheet.append([])
                summary_sheet.append(["Debug Information"])
                for i, result in enumerate(batch_results):
                    summary_sheet.append([f"Image {i+1} Status", "Success" if result["success"] else "Failed"])
                    if result.get("data"):
                        summary_sheet.append([f"Image {i+1} Data Length", str(len(result["data"]))])
                        summary_sheet.append([f"Image {i+1} Data Preview", result["data"][:100]])
            
            # Save to bytes
            output = io.BytesIO()
            workbook.save(output)
            output.seek(0)
            
            return output.getvalue()
            
        except Exception as e:
            logger.error(f"Failed to create consolidated XLSX: {str(e)}")
            raise FileProcessingError(f"Failed to create consolidated Excel file: {str(e)}")
    
    def _create_concatenated_xlsx(self, batch_results: List[Dict[str, Any]]) -> bytes:
        """
        Create a single XLSX file with concatenated data.
        
        Args:
            batch_results: List of batch processing results
            
        Returns:
            XLSX file as bytes
        """
        try:
            all_dataframes = []
            
            for i, result in enumerate(batch_results):
                if result["success"] and result["data"]:
                    try:
                        df = pd.read_csv(io.StringIO(result["data"]))
                        df = df.fillna("")
                        
                        # Add source column
                        df.insert(0, "Source", f"Image_{i+1}")
                        all_dataframes.append(df)
                        
                    except Exception as e:
                        logger.warning(f"Failed to process result {i} for concatenation: {str(e)}")
            
            if not all_dataframes:
                raise FileProcessingError("No valid data to concatenate")
            
            # Concatenate all dataframes
            combined_df = pd.concat(all_dataframes, ignore_index=True, sort=False)
            
            # Convert to XLSX
            return self._dataframe_to_xlsx_bytes(combined_df, "Combined_Data")
            
        except Exception as e:
            logger.error(f"Failed to create concatenated XLSX: {str(e)}")
            raise FileProcessingError(f"Failed to create concatenated Excel file: {str(e)}")
    
    def _create_separate_xlsx(self, batch_results: List[Dict[str, Any]]) -> bytes:
        """
        Create a ZIP file containing separate XLSX files.
        Note: For now, returns the first successful result as XLSX.
        TODO: Implement ZIP creation for multiple files.
        
        Args:
            batch_results: List of batch processing results
            
        Returns:
            XLSX file as bytes
        """
        # For now, return the first successful result
        # In a full implementation, this would create a ZIP file
        for i, result in enumerate(batch_results):
            if result["success"] and result["data"]:
                return self.csv_to_xlsx(result["data"], f"Table_{i+1}")
        
        raise FileProcessingError("No successful results to process")
    
    def _dataframe_to_xlsx_bytes(self, df: pd.DataFrame, sheet_name: str) -> bytes:
        """
        Convert pandas DataFrame to XLSX bytes with styling.
        
        Args:
            df: DataFrame to convert
            sheet_name: Name for the Excel sheet
            
        Returns:
            XLSX file as bytes
        """
        output = io.BytesIO()
        
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            
            # Get the workbook and worksheet
            workbook = writer.book
            worksheet = writer.sheets[sheet_name]
            
            # Style the header row
            self._style_header_row(worksheet)
            
            # Auto-adjust column widths
            self._auto_adjust_columns(worksheet, df)
        
        output.seek(0)
        return output.getvalue()

    def bank_statement_to_xlsx(self, statement_data: Dict[str, Any], sheet_name: str = "Statement") -> bytes:
        """
        Create a bank statement workbook that preserves what the user sees:
        statement text first, transaction table underneath, plus review sheets.
        """
        try:
            workbook = Workbook()
            statement_sheet = workbook.active
            statement_sheet.title = sheet_name[:31] or "Statement"

            summary = statement_data.get("summary") or {
                "account_holder": statement_data.get("account_holder", ""),
                "bank_name": statement_data.get("bank_name", ""),
                "account_number": statement_data.get("account_number", ""),
                "statement_period": statement_data.get("period", ""),
                "currency": statement_data.get("currency", ""),
                "opening_balance": statement_data.get("opening_balance", ""),
                "closing_balance": statement_data.get("closing_balance", ""),
            }
            summary_rows = self._normalize_rows(summary, ["Field", "Value"])
            transaction_rows = [["Page", "Date", "Description", "Reference", "Debit", "Credit", "Balance"]]
            for transaction in statement_data.get("transactions") or []:
                if isinstance(transaction, dict):
                    transaction_rows.append([
                        transaction.get("page", ""),
                        transaction.get("date", ""),
                        transaction.get("description", ""),
                        transaction.get("reference", ""),
                        transaction.get("debit", ""),
                        transaction.get("credit", ""),
                        transaction.get("balance", ""),
                    ])
            review_rows = [["Code", "Field", "Note"]]
            for flag in statement_data.get("review_flags") or statement_data.get("review") or []:
                if isinstance(flag, dict):
                    review_rows.append([
                        flag.get("code", "statement_review_note"),
                        flag.get("area", ""),
                        flag.get("note", ""),
                    ])
            raw_text = self._normalize_raw_text(statement_data.get("raw_text"))

            statement_sheet.append(["Detected Statement Text"])
            statement_sheet.merge_cells(start_row=2, start_column=1, end_row=8, end_column=8)
            statement_sheet.cell(row=2, column=1, value=raw_text or "No non-table text detected.")
            statement_sheet.cell(row=2, column=1).alignment = Alignment(wrap_text=True, vertical="top")
            statement_sheet.append([])
            table_title_row = statement_sheet.max_row + 1
            statement_sheet.append(["Detected Transactions"])
            for row in transaction_rows:
                statement_sheet.append(row)
            self._style_header_row(statement_sheet, table_title_row + 1)
            self._style_statement_sheet(statement_sheet)

            summary_sheet = workbook.create_sheet("Summary")
            for row in summary_rows:
                summary_sheet.append(row)
            self._style_header_row(summary_sheet)
            self._auto_fit_openpyxl_columns(summary_sheet)

            transactions_sheet = workbook.create_sheet("Transactions")
            for row in transaction_rows:
                transactions_sheet.append(row)
            self._style_header_row(transactions_sheet)
            self._auto_fit_openpyxl_columns(transactions_sheet)

            raw_sheet = workbook.create_sheet("Raw Text")
            raw_sheet.append(["Visible Text"])
            raw_sheet.append([raw_text or "No readable text detected."])
            raw_sheet["A2"].alignment = Alignment(wrap_text=True, vertical="top")
            raw_sheet.column_dimensions["A"].width = 120
            self._style_header_row(raw_sheet)

            review_sheet = workbook.create_sheet("Review Flags")
            for row in review_rows:
                review_sheet.append(row)
            self._style_header_row(review_sheet)
            self._auto_fit_openpyxl_columns(review_sheet)

            output = io.BytesIO()
            workbook.save(output)
            output.seek(0)
            return output.getvalue()
        except Exception as e:
            logger.error(f"Bank statement workbook error: {str(e)}")
            return self._create_raw_text_xlsx(json.dumps(statement_data, ensure_ascii=False, indent=2), sheet_name)

    def bank_statement_transactions_to_csv(self, statement_data: Dict[str, Any]) -> bytes:
        """Create a clean CSV export of visible bank statement transactions."""
        output = io.StringIO(newline="")
        writer = csv.writer(output)
        columns = [
            ("page", "Page"),
            ("date", "Date"),
            ("description", "Description"),
            ("reference", "Reference"),
            ("debit", "Debit"),
            ("credit", "Credit"),
            ("balance", "Balance"),
        ]
        writer.writerow([label for _, label in columns])
        for row in statement_data.get("transactions") or []:
            if isinstance(row, dict):
                writer.writerow([row.get(key, "") for key, _ in columns])
        return output.getvalue().encode("utf-8-sig")

    def notes_text_to_txt(self, notes_data: Dict[str, Any]) -> bytes:
        """Return readable narrative notes text without inventing table structure."""
        return str(notes_data.get("readable_text") or "").encode("utf-8")

    def notes_tables_to_xlsx(self, notes_data: Dict[str, Any]) -> bytes:
        """Create a workbook from tables explicitly detected in a notes page."""
        workbook = Workbook()
        default_sheet = workbook.active
        tables = [table for table in notes_data.get("tables") or [] if isinstance(table, dict)]
        used_names = set()
        for index, table in enumerate(tables):
            raw_title = table.get("title") or f"Table {index + 1}"
            title = "".join(char if char not in r'[]:*?/\\' else "_" for char in str(raw_title))[:31] or f"Table {index + 1}"
            candidate = title
            counter = 2
            while candidate in used_names:
                suffix = f"_{counter}"
                candidate = f"{title[:31-len(suffix)]}{suffix}"
                counter += 1
            used_names.add(candidate)
            sheet = default_sheet if index == 0 else workbook.create_sheet()
            sheet.title = candidate
            columns = table.get("columns") or []
            if columns:
                sheet.append(columns)
                self._style_header_row(sheet)
            for row in table.get("rows") or []:
                if isinstance(row, list):
                    sheet.append(row)
            self._auto_fit_openpyxl_columns(sheet)
        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)
        return output.getvalue()

    def notes_tables_to_csv(self, notes_data: Dict[str, Any]) -> bytes:
        """Create CSV rows only from tables explicitly detected in notes."""
        output = io.StringIO(newline="")
        writer = csv.writer(output)
        tables = [table for table in notes_data.get("tables") or [] if isinstance(table, dict)]
        for index, table in enumerate(tables):
            if index:
                writer.writerow([])
            if len(tables) > 1:
                writer.writerow([table.get("title") or f"Table {index + 1}"])
            columns = table.get("columns") or []
            if columns:
                writer.writerow(columns)
            for row in table.get("rows") or []:
                if isinstance(row, list):
                    writer.writerow(row)
        return output.getvalue().encode("utf-8-sig")

    def invoice_to_xlsx(self, invoice_data: Dict[str, Any]) -> bytes:
        """Create a review-ready invoice workbook with summary and line-item sheets."""
        try:
            workbook = Workbook()
            summary_sheet = workbook.active
            summary_sheet.title = "Invoice Summary"
            summary_fields = [
                ("Vendor name", "vendor_name"),
                ("Supplier tax/VAT ID", "supplier_tax_vat_id"),
                ("Invoice number", "invoice_number"),
                ("Invoice date", "invoice_date"),
                ("Due date", "due_date"),
                ("PO/reference", "po_reference"),
                ("Subtotal", "subtotal"),
                ("Tax/VAT amount", "tax_vat_amount"),
                ("Total", "total"),
                ("Currency", "currency"),
            ]
            summary_sheet.append(["Field", "Value"])
            for label, key in summary_fields:
                summary_sheet.append([label, invoice_data.get(key, "")])
            self._style_header_row(summary_sheet)
            self._auto_fit_openpyxl_columns(summary_sheet)

            items_sheet = workbook.create_sheet("Line Items")
            item_fields = [
                ("Description", "description"),
                ("Quantity", "quantity"),
                ("Unit price", "unit_price"),
                ("Tax/rate", "tax_rate"),
                ("Line total", "line_total"),
            ]
            items_sheet.append([label for label, _ in item_fields])
            for item in invoice_data.get("line_items") or []:
                if isinstance(item, dict):
                    items_sheet.append([item.get(key, "") for _, key in item_fields])
            self._style_header_row(items_sheet)
            self._auto_fit_openpyxl_columns(items_sheet)

            review_sheet = workbook.create_sheet("Review Flags")
            review_sheet.append(["Code", "Field", "Note"])
            for flag in invoice_data.get("review_flags") or []:
                if isinstance(flag, dict):
                    review_sheet.append([
                        flag.get("code", ""),
                        flag.get("area", ""),
                        flag.get("note", ""),
                    ])
            self._style_header_row(review_sheet)
            self._auto_fit_openpyxl_columns(review_sheet)

            output = io.BytesIO()
            workbook.save(output)
            output.seek(0)
            return output.getvalue()
        except Exception as e:
            logger.error(f"Invoice workbook error: {str(e)}")
            return self._create_raw_text_xlsx(json.dumps(invoice_data, ensure_ascii=False, indent=2), "Invoice")

    def invoice_line_items_to_csv(self, invoice_data: Dict[str, Any]) -> bytes:
        """Create a CSV export containing invoice identity fields and line items."""
        output = io.StringIO(newline="")
        writer = csv.writer(output)
        summary_columns = [
            ("vendor_name", "Vendor name"),
            ("supplier_tax_vat_id", "Supplier tax/VAT ID"),
            ("invoice_number", "Invoice number"),
            ("invoice_date", "Invoice date"),
            ("due_date", "Due date"),
            ("po_reference", "PO/reference"),
            ("subtotal", "Subtotal"),
            ("tax_vat_amount", "Tax/VAT amount"),
            ("total", "Total"),
            ("currency", "Currency"),
        ]
        item_columns = [
            ("description", "Description"),
            ("quantity", "Quantity"),
            ("unit_price", "Unit price"),
            ("tax_rate", "Tax/rate"),
            ("line_total", "Line total"),
        ]
        writer.writerow([label for _, label in summary_columns + item_columns])
        items = [item for item in invoice_data.get("line_items") or [] if isinstance(item, dict)] or [{}]
        for item in items:
            writer.writerow(
                [invoice_data.get(key, "") for key, _ in summary_columns]
                + [item.get(key, "") for key, _ in item_columns]
            )
        return output.getvalue().encode("utf-8-sig")

    def receipt_to_xlsx(self, receipt_data: Dict[str, Any]) -> bytes:
        """Create an expense-review workbook from visible receipt fields."""
        try:
            workbook = Workbook()
            summary_sheet = workbook.active
            summary_sheet.title = "Receipt Summary"
            summary_fields = [
                ("Merchant", "merchant"),
                ("Date", "date"),
                ("Payment method", "payment_method"),
                ("Currency", "currency"),
                ("Subtotal", "subtotal"),
                ("Tax/VAT amount", "tax_vat_amount"),
                ("Discount", "discount"),
                ("Tip", "tip"),
                ("Total", "total"),
            ]
            summary_sheet.append(["Field", "Value"])
            for label, key in summary_fields:
                summary_sheet.append([label, receipt_data.get(key, "")])
            self._style_header_row(summary_sheet)
            self._auto_fit_openpyxl_columns(summary_sheet)

            items_sheet = workbook.create_sheet("Line Items")
            item_fields = [
                ("Description", "description"),
                ("Quantity", "quantity"),
                ("Unit price", "unit_price"),
                ("Tax/rate", "tax_rate"),
                ("Line total", "line_total"),
            ]
            items_sheet.append([label for label, _ in item_fields])
            for item in receipt_data.get("line_items") or []:
                if isinstance(item, dict):
                    items_sheet.append([item.get(key, "") for _, key in item_fields])
            self._style_header_row(items_sheet)
            self._auto_fit_openpyxl_columns(items_sheet)

            review_sheet = workbook.create_sheet("Review Flags")
            review_sheet.append(["Code", "Field", "Note"])
            for flag in receipt_data.get("review_flags") or []:
                if isinstance(flag, dict):
                    review_sheet.append([
                        flag.get("code", ""),
                        flag.get("area", ""),
                        flag.get("note", ""),
                    ])
            self._style_header_row(review_sheet)
            self._auto_fit_openpyxl_columns(review_sheet)

            output = io.BytesIO()
            workbook.save(output)
            output.seek(0)
            return output.getvalue()
        except Exception as e:
            logger.error(f"Receipt workbook error: {str(e)}")
            return self._create_raw_text_xlsx(json.dumps(receipt_data, ensure_ascii=False, indent=2), "Receipt")

    def receipt_line_items_to_csv(self, receipt_data: Dict[str, Any]) -> bytes:
        """Create an expense-review CSV containing receipt fields and visible items."""
        output = io.StringIO(newline="")
        writer = csv.writer(output)
        summary_columns = [
            ("merchant", "Merchant"),
            ("date", "Date"),
            ("payment_method", "Payment method"),
            ("currency", "Currency"),
            ("subtotal", "Subtotal"),
            ("tax_vat_amount", "Tax/VAT amount"),
            ("discount", "Discount"),
            ("tip", "Tip"),
            ("total", "Total"),
        ]
        item_columns = [
            ("description", "Description"),
            ("quantity", "Quantity"),
            ("unit_price", "Unit price"),
            ("tax_rate", "Tax/rate"),
            ("line_total", "Line total"),
        ]
        writer.writerow([label for _, label in summary_columns + item_columns])
        items = [item for item in receipt_data.get("line_items") or [] if isinstance(item, dict)] or [{}]
        for item in items:
            writer.writerow(
                [receipt_data.get(key, "") for key, _ in summary_columns]
                + [item.get(key, "") for key, _ in item_columns]
            )
        return output.getvalue().encode("utf-8-sig")

    def export_reviewed_document(
        self,
        document: Dict[str, Any],
        requested_format: str = "xlsx",
    ) -> Dict[str, Any]:
        """Render one durable document from reviewed values rather than raw OCR output."""
        mode, reviewed_data = self._merge_reviewed_document_data(document)
        output_format = str(requested_format or "xlsx").lower()
        if output_format not in {"xlsx", "csv", "txt"}:
            output_format = "xlsx"

        if mode == "invoice":
            if output_format == "csv":
                content = self.invoice_line_items_to_csv(reviewed_data)
                extension = "csv"
                content_type = "text/csv; charset=utf-8"
            else:
                content = self.invoice_to_xlsx(reviewed_data)
                extension = "xlsx"
                content_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        elif mode == "receipt":
            if output_format == "csv":
                content = self.receipt_line_items_to_csv(reviewed_data)
                extension = "csv"
                content_type = "text/csv; charset=utf-8"
            else:
                content = self.receipt_to_xlsx(reviewed_data)
                extension = "xlsx"
                content_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        elif mode == "bank_statement":
            if output_format == "csv":
                content = self.bank_statement_transactions_to_csv(reviewed_data)
                extension = "csv"
                content_type = "text/csv; charset=utf-8"
            else:
                content = self.bank_statement_to_xlsx(reviewed_data, "Statement")
                extension = "xlsx"
                content_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        elif mode == "notes":
            has_tables = bool(reviewed_data.get("tables"))
            if output_format == "csv" and has_tables:
                content = self.notes_tables_to_csv(reviewed_data)
                extension = "csv"
                content_type = "text/csv; charset=utf-8"
            elif output_format == "xlsx" and has_tables:
                content = self.notes_tables_to_xlsx(reviewed_data)
                extension = "xlsx"
                content_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            else:
                content = self.notes_text_to_txt(reviewed_data)
                extension = "txt"
                content_type = "text/plain; charset=utf-8"
        else:
            csv_content = str(reviewed_data.get("csv") or "")
            if output_format == "csv":
                content = self.table_to_csv(csv_content)
                extension = "csv"
                content_type = "text/csv; charset=utf-8"
            else:
                content = self.csv_to_xlsx(csv_content, "Table")
                extension = "xlsx"
                content_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

        original_name = document.get("original_filename") or "document"
        filename = f"{self._safe_export_stem(original_name)}_reviewed.{extension}"
        return {
            "content": content,
            "filename": filename,
            "content_type": content_type,
            "mode": mode,
        }

    def _merge_reviewed_document_data(self, document: Dict[str, Any]) -> tuple[str, Dict[str, Any]]:
        """Combine reviewed page/unit values into the logical document export shape."""
        mode = str(
            document.get("resolved_mode")
            or document.get("detected_mode")
            or document.get("selected_mode")
            or "table"
        )
        extractions = sorted(
            document.get("extractions") or [],
            key=self._extraction_page_order,
        )
        reviewed_units = [
            (extraction, self._reviewed_extraction_data(extraction))
            for extraction in extractions
            if isinstance(extraction, dict)
        ]
        payloads = [payload for _, payload in reviewed_units if payload]

        if mode == "invoice_receipt":
            mode = "invoice" if any(payload.get("invoice_number") for payload in payloads) else "receipt"

        if mode in {"invoice", "receipt", "bank_statement"}:
            data: Dict[str, Any] = {}
            list_key = "transactions" if mode == "bank_statement" else "line_items"
            data[list_key] = []
            data["review_flags"] = []
            for extraction, payload in reviewed_units:
                for key, value in payload.items():
                    if key not in {list_key, "review_flags"} and value not in (None, "", [], {}):
                        if data.get(key) in (None, "", [], {}):
                            data[key] = value
                data[list_key].extend(
                    item for item in payload.get(list_key) or [] if isinstance(item, dict)
                )
                data["review_flags"].extend(
                    flag for flag in payload.get("review_flags") or [] if isinstance(flag, dict)
                )
                data["review_flags"].extend(
                    flag for flag in extraction.get("validation_flags") or [] if isinstance(flag, dict)
                )
            return mode, data

        if mode == "notes":
            readable_pages = [
                str(payload.get("readable_text") or "").strip()
                for payload in payloads
                if str(payload.get("readable_text") or "").strip()
            ]
            data = {"readable_text": "\n\n".join(readable_pages), "tables": [], "review_flags": []}
            for extraction, payload in reviewed_units:
                data["tables"].extend(table for table in payload.get("tables") or [] if isinstance(table, dict))
                data["review_flags"].extend(
                    flag for flag in payload.get("review_flags") or [] if isinstance(flag, dict)
                )
                data["review_flags"].extend(
                    flag for flag in extraction.get("validation_flags") or [] if isinstance(flag, dict)
                )
            return mode, data

        merged_grid: List[List[Any]] = []
        for payload in payloads:
            grid = payload.get("review_grid")
            if not isinstance(grid, list):
                csv_content = payload.get("csv")
                grid = list(csv.reader(io.StringIO(csv_content))) if isinstance(csv_content, str) else []
            clean_grid = [list(row) for row in grid if isinstance(row, list)]
            if merged_grid and clean_grid and clean_grid[0] == merged_grid[0]:
                clean_grid = clean_grid[1:]
            merged_grid.extend(clean_grid)
        output = io.StringIO(newline="")
        csv.writer(output).writerows(merged_grid)
        return "table", {"review_grid": merged_grid, "csv": output.getvalue()}

    @staticmethod
    def _reviewed_extraction_data(extraction: Dict[str, Any]) -> Dict[str, Any]:
        reviewed = extraction.get("reviewed_data")
        if isinstance(reviewed, dict) and reviewed:
            return reviewed
        structured = extraction.get("raw_structured_data") or extraction.get("structured_data")
        return structured if isinstance(structured, dict) else {}

    @staticmethod
    def _extraction_page_order(extraction: Dict[str, Any]) -> tuple[int, str]:
        metadata = extraction.get("metadata") if isinstance(extraction.get("metadata"), dict) else {}
        source_page = extraction.get("source_page") or metadata.get("source_page") or 0
        try:
            page = int(source_page)
        except (TypeError, ValueError):
            page = 0
        return page, str(extraction.get("processing_unit_id") or "")

    @staticmethod
    def _safe_export_stem(filename: str) -> str:
        stem = os.path.splitext(os.path.basename(str(filename or "document")))[0]
        cleaned = re.sub(r"[^A-Za-z0-9._-]+", "_", stem).strip("._-")
        return cleaned or "document"

    def _normalize_rows(self, value: Any, fallback_header: List[str]) -> List[List[Any]]:
        if not value:
            return [fallback_header]
        if isinstance(value, list):
            if all(isinstance(row, dict) for row in value):
                keys: List[str] = []
                for row in value:
                    for key in row.keys():
                        if key not in keys:
                            keys.append(key)
                header = keys or fallback_header
                return [header] + [[row.get(key, "") for key in header] for row in value]
            if all(isinstance(row, list) for row in value):
                return value or [fallback_header]
        if isinstance(value, dict):
            return [fallback_header] + [[key, val] for key, val in value.items()]
        return [fallback_header, ["Detected", str(value)]]

    def _normalize_raw_text(self, value: Any) -> str:
        if isinstance(value, str):
            return value.strip()
        if isinstance(value, list):
            return "\n\n".join(str(item).strip() for item in value if str(item).strip())
        if isinstance(value, dict):
            return "\n".join(f"{key}: {val}" for key, val in value.items())
        return str(value or "").strip()
    
    def _style_header_row(self, worksheet, row_number: int = 1) -> None:
        """
        Style the header row of a worksheet.
        
        Args:
            worksheet: openpyxl worksheet object
        """
        if worksheet.max_row >= row_number:
            for cell in worksheet[row_number]:
                cell.font = self.header_font
                cell.fill = self.header_fill
                cell.alignment = self.alignment

    def _style_statement_sheet(self, worksheet) -> None:
        worksheet["A1"].font = self.header_font
        worksheet["A1"].fill = self.header_fill
        worksheet["A1"].alignment = self.alignment
        worksheet.row_dimensions[2].height = 130
        for column in range(1, 9):
            worksheet.column_dimensions[chr(64 + column)].width = 18
        for row in worksheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")

    def _auto_fit_openpyxl_columns(self, worksheet) -> None:
        for column_cells in worksheet.columns:
            column_letter = column_cells[0].column_letter
            max_length = max((len(str(cell.value or "")) for cell in column_cells), default=10)
            worksheet.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 48)
    
    def _auto_adjust_columns(self, worksheet, df: pd.DataFrame) -> None:
        """
        Auto-adjust column widths based on content.
        
        Args:
            worksheet: openpyxl worksheet object
            df: DataFrame for width calculation
        """
        for i, column in enumerate(df.columns):
            # Calculate max width needed
            max_length = max(
                len(str(column)),  # Header length
                df[column].astype(str).map(len).max() if not df.empty else 0  # Content length
            )
            
            # Set column width (with some padding)
            adjusted_width = min(max_length + 2, 50)  # Max width of 50
            worksheet.column_dimensions[chr(65 + i)].width = adjusted_width
    
    def _create_empty_xlsx_with_message(self, message: str, sheet_name: str) -> bytes:
        """
        Create an Excel file with a message when no data is available.
        
        Args:
            message: Message to display in the Excel file
            sheet_name: Name for the Excel sheet
            
        Returns:
            XLSX file as bytes
        """
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = sheet_name
        
        # Add message
        worksheet['A1'] = "Status"
        worksheet['B1'] = message
        
        # Add some basic styling
        worksheet['A1'].font = self.header_font
        worksheet['A1'].fill = self.header_fill
        worksheet['B1'].font = Font(color="FF0000")  # Red text
        
        # Adjust column widths
        worksheet.column_dimensions['A'].width = 15
        worksheet.column_dimensions['B'].width = 50
        
        # Save to bytes
        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)
        return output.getvalue()
    
    def _create_raw_text_xlsx(self, content: str, sheet_name: str) -> bytes:
        """
        Create an Excel file with raw text content when CSV parsing fails.
        
        Args:
            content: Raw text content to include
            sheet_name: Name for the Excel sheet
            
        Returns:
            XLSX file as bytes
        """
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = sheet_name
        
        # Add headers
        worksheet['A1'] = "Raw OCR Output"
        worksheet['A1'].font = self.header_font
        worksheet['A1'].fill = self.header_fill
        
        # Split content into lines and add to Excel
        lines = content.split('\n')
        for i, line in enumerate(lines, start=2):
            if line.strip():  # Only add non-empty lines
                worksheet[f'A{i}'] = line.strip()
        
        # Adjust column width
        worksheet.column_dimensions['A'].width = 80
        
        # Save to bytes
        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)
        return output.getvalue()

    def create_multi_sheet_xlsx(self, results: List[Dict[str, Any]]) -> bytes:
        """
        Create an XLSX file with multiple sheets from batch results.
        
        Args:
            results: List of result dictionaries with 'sheet_name', 'csv_data', and optionally 'error'
            
        Returns:
            XLSX file as bytes
        """
        try:
            workbook = Workbook()
            # Remove default sheet
            workbook.remove(workbook.active)
            
            for i, result in enumerate(results):
                sheet_name = result.get('sheet_name', f'Sheet_{i+1}')
                # Clean sheet name (Excel has restrictions)
                sheet_name = sheet_name.replace('/', '_').replace('\\', '_')[:31]
                
                worksheet = workbook.create_sheet(title=sheet_name)
                
                if result.get('error'):
                    # Add error information
                    worksheet['A1'] = "Error"
                    worksheet['A1'].font = self.header_font
                    worksheet['A1'].fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                    worksheet['A2'] = f"Failed to process: {result['error']}"
                else:
                    csv_data = result.get('csv_data', '')
                    if csv_data and csv_data.strip():
                        # Parse CSV and add to worksheet
                        try:
                            df = pd.read_csv(io.StringIO(csv_data))
                            if not df.empty:
                                # Add headers
                                for col_num, column_name in enumerate(df.columns, 1):
                                    cell = worksheet.cell(row=1, column=col_num, value=column_name)
                                    cell.font = self.header_font
                                    cell.fill = self.header_fill
                                    cell.alignment = self.alignment
                                
                                # Add data
                                for row_num, row_data in df.iterrows():
                                    for col_num, value in enumerate(row_data, 1):
                                        worksheet.cell(row=row_num + 2, column=col_num, value=value)
                                
                                # Auto-adjust column widths
                                for column in worksheet.columns:
                                    max_length = 0
                                    column_letter = column[0].column_letter
                                    for cell in column:
                                        try:
                                            if len(str(cell.value)) > max_length:
                                                max_length = len(str(cell.value))
                                        except:
                                            pass
                                    adjusted_width = min(max_length + 2, 50)
                                    worksheet.column_dimensions[column_letter].width = adjusted_width
                            else:
                                worksheet['A1'] = "No data extracted"
                        except Exception as e:
                            worksheet['A1'] = f"Parse error: {str(e)}"
                    else:
                        worksheet['A1'] = "No data extracted from image"
            
            # If no sheets were created, add a default one
            if not workbook.worksheets:
                worksheet = workbook.create_sheet(title="Results")
                worksheet['A1'] = "No results to display"
            
            # Save to bytes
            output = io.BytesIO()
            workbook.save(output)
            output.seek(0)
            return output.getvalue()
            
        except Exception as e:
            logger.error(f"Error creating multi-sheet XLSX: {str(e)}")
            # Return a basic error sheet
            return self._create_empty_xlsx_with_message(f"Error creating results: {str(e)}", "Error")

    def create_concatenated_xlsx(self, results: List[Dict[str, Any]]) -> bytes:
        """
        Create an XLSX file with all results concatenated in a single sheet.
        
        Args:
            results: List of result dictionaries with 'sheet_name', 'csv_data', and optionally 'error'
            
        Returns:
            XLSX file as bytes
        """
        try:
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "Concatenated_Results"
            
            current_row = 1
            
            for i, result in enumerate(results):
                sheet_name = result.get('sheet_name', f'Image_{i+1}')
                
                # Add section header
                header_cell = worksheet.cell(row=current_row, column=1, value=f"=== {sheet_name} ===")
                header_cell.font = Font(bold=True, size=14)
                header_cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
                current_row += 1
                
                if result.get('error'):
                    error_cell = worksheet.cell(row=current_row, column=1, value=f"Error: {result['error']}")
                    error_cell.font = Font(color="FF0000")
                    current_row += 2
                else:
                    csv_data = result.get('csv_data', '')
                    if csv_data and csv_data.strip():
                        try:
                            df = pd.read_csv(io.StringIO(csv_data))
                            if not df.empty:
                                # Add headers
                                for col_num, column_name in enumerate(df.columns, 1):
                                    cell = worksheet.cell(row=current_row, column=col_num, value=column_name)
                                    cell.font = self.header_font
                                    cell.fill = self.header_fill
                                current_row += 1
                                
                                # Add data rows
                                for _, row_data in df.iterrows():
                                    for col_num, value in enumerate(row_data, 1):
                                        worksheet.cell(row=current_row, column=col_num, value=value)
                                    current_row += 1
                            else:
                                worksheet.cell(row=current_row, column=1, value="No data extracted")
                                current_row += 1
                        except Exception as e:
                            worksheet.cell(row=current_row, column=1, value=f"Parse error: {str(e)}")
                            current_row += 1
                    else:
                        worksheet.cell(row=current_row, column=1, value="No data extracted")
                        current_row += 1
                
                # Add spacing between sections
                current_row += 1
            
            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            # Save to bytes
            output = io.BytesIO()
            workbook.save(output)
            output.seek(0)
            return output.getvalue()
            
        except Exception as e:
            logger.error(f"Error creating concatenated XLSX: {str(e)}")
            return self._create_empty_xlsx_with_message(f"Error creating results: {str(e)}", "Error")

    def _create_empty_xlsx_with_message(self, message: str, sheet_name: str = "Results") -> bytes:
        """
        Create an empty XLSX file with a single message.
        
        Args:
            message: Message to display
            sheet_name: Name of the sheet
            
        Returns:
            XLSX file as bytes
        """
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = sheet_name
        
        cell = worksheet['A1']
        cell.value = message
        cell.font = self.header_font
        cell.fill = PatternFill(start_color="FFAAAA", end_color="FFAAAA", fill_type="solid")
        
        # Save to bytes
        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)
        return output.getvalue()


# Global Excel service instance
excel_service = ExcelService()
